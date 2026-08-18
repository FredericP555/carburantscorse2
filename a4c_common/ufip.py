#!/usr/bin/env python3
"""Client for the public UFIP / Énergies et Mobilités custom-value export."""
from __future__ import annotations

import io
from datetime import date

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

UFIP_CUSTOM_URL = "https://valeurs.ufip.fr/datas/custom"
USER_AGENT = "A4C-observatoires/2.0 (+public-data research)"
GAZOLE_HEADER_PREFIX = "GAZOLE (Rotterdam)"


def _format_date(value: date) -> str:
    return value.strftime("%d/%m/%Y")


def parse_rotterdam_gazole_xlsx(raw: bytes) -> pd.DataFrame:
    """Parse the two-column UFIP export into date/value rows."""
    if not raw.startswith(b"PK"):
        raise ValueError("UFIP response is not an XLSX ZIP container")
    wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
    if not wb.worksheets:
        raise ValueError("UFIP workbook has no worksheet")
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("UFIP workbook is empty")
    header = [str(v).strip() if v is not None else "" for v in rows[0]]
    try:
        date_col = header.index("Date")
    except ValueError as exc:
        raise ValueError(f"UFIP workbook has no Date column: {header}") from exc
    fuel_col = next((i for i, value in enumerate(header) if value.startswith(GAZOLE_HEADER_PREFIX)), None)
    if fuel_col is None:
        raise ValueError(f"UFIP workbook has no Rotterdam Gazole column: {header}")

    parsed = []
    for row in rows[1:]:
        if date_col >= len(row) or fuel_col >= len(row):
            continue
        raw_date, raw_value = row[date_col], row[fuel_col]
        if raw_date is None or raw_value is None:
            continue
        if hasattr(raw_date, "date"):
            d = raw_date.date()
        elif isinstance(raw_date, date):
            d = raw_date
        else:
            d = pd.to_datetime(raw_date, dayfirst=True).date()
        parsed.append((d, float(raw_value)))
    df = pd.DataFrame(parsed, columns=["date", "rotterdam_eur_l"])
    if not df.empty:
        df = df.sort_values("date").drop_duplicates("date", keep="last").reset_index(drop=True)
    return df


def fetch_rotterdam_gazole(
    start_date: date,
    end_date: date,
    *,
    session: requests.Session | None = None,
    timeout: int = 90,
) -> pd.DataFrame:
    """Download the UFIP Rotterdam Gazole series for a custom period.

    Audit performed on 18 Aug 2026 established the public mechanism: GET the form
    to obtain ``ufp_token`` and the session cookie, then POST the same URL with
    ``day_from``, ``day_to`` and ``cotations[gazole]=on``.
    """
    if end_date < start_date:
        raise ValueError("end_date must be >= start_date")
    own_session = session is None
    s = session or requests.Session()
    s.headers.setdefault("User-Agent", USER_AGENT)
    try:
        first = s.get(UFIP_CUSTOM_URL, timeout=timeout)
        first.raise_for_status()
        soup = BeautifulSoup(first.text, "html.parser")
        token = soup.select_one('input[name="ufp_token"]')
        if token is None or not token.get("value"):
            raise RuntimeError("UFIP ufp_token not found in custom export form")
        payload = {
            "ufp_token": token["value"],
            "day_from": _format_date(start_date),
            "day_to": _format_date(end_date),
            "cotations[gazole]": "on",
        }
        response = s.post(UFIP_CUSTOM_URL, data=payload, timeout=timeout, allow_redirects=True)
        response.raise_for_status()
        return parse_rotterdam_gazole_xlsx(response.content)
    finally:
        if own_session:
            s.close()


def expand_daily(observations: pd.DataFrame, start_date: date, end_date: date) -> pd.DataFrame:
    """Forward-fill weekends/holidays from the last UFIP observation."""
    if end_date < start_date:
        raise ValueError("end_date must be >= start_date")
    calendar = pd.DataFrame({"date": pd.date_range(start_date, end_date, freq="D").date})
    source = observations.copy()
    if not source.empty:
        source["date"] = pd.to_datetime(source["date"]).dt.date
        source = source.sort_values("date").drop_duplicates("date", keep="last")
        source["rotterdam_observed"] = True
    merged = calendar.merge(source, on="date", how="left")
    if "rotterdam_observed" not in merged.columns:
        merged["rotterdam_observed"] = False
    else:
        merged["rotterdam_observed"] = merged["rotterdam_observed"].eq(True)
    merged["rotterdam_eur_l"] = pd.to_numeric(merged.get("rotterdam_eur_l"), errors="coerce").ffill()
    merged["rotterdam_carried"] = merged["rotterdam_eur_l"].notna() & ~merged["rotterdam_observed"]
    return merged
